# Main script

def main(stock1,stock2,stock3,stock4,stock5,stock6, timePeriod):
    import json
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Alignment, PatternFill
    from getJSON import getJSON as g
    from io import BytesIO
    from addSheet import getArray
    from openpyxl.chart import LineChart, Reference, BarChart, AreaChart
    from getLogo import download_image, save_image_to_file
    import datetime as dt
    from createChart import createChart
    from openpyxl.drawing.text import Font
    
    # Excel Declarations
    wb = Workbook()
    ws = wb.active
    apiHost = "yahoo-finance-api-data.p.rapidapi.com"
    image = Image
    ws._image = image
    symbols = [stock1,stock2,stock3,stock4,stock5,stock6]
    symbolsUsed = ['','','','','',''] 

    # Ignore blank stock textboxes
    index = 0
    for s in range(6):
        if symbols[s] is not '':
            symbolsUsed[index] = symbols[s]
            index += 1          
    symbols = symbolsUsed
    numNonBlank = index

    # Get logo pictures
    logoDestination = ['B1','C1','D1','E1','F1','G1']
    stockDestination =  ['B25','C25','D25','E25','F25','G25']
    ws.row_dimensions[1].height = 100
    ws.row_dimensions[13].height = 30
    ws.row_dimensions[19].height = 30

    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20

    # Loop thru Symbols
    validCode = [True, True, True, True, True, True]
    for c in range(index): #used to be 6
        print("c=" + str(c))
        # Get company info from 'symbol-profile' path of Yahoo Finance API
        url = "https://yahoo-finance-api-data.p.rapidapi.com/summary/symbol-profile?symbol=" + symbols[c]
        response = g(url,apiHost)
        companyAttributes = ['address1','city','state','zip','country','phone','website','industry','sector','fullTimeEmployees']
        companyAttributeValues = ['address1','city','state','zip','country','phone','website','industry','sector','fullTimeEmployees']
        counter = 12
        #print("c=" + str(c) + " and response = " + str(response['success']))
        if response['success'] == False:
            validCode[c] = False
            ws.cell(3,c+2,symbols[c])
            ws.cell(2,c+2,'Symbol not found' )

            continue
        for element in companyAttributes:
            counter+=1
            if element in response['data']:
                value = response['data'][element]
                if element == 'website':
                        print (value)

                        # Get logo pictures from clearbit
                        image_url = 'https://logo.clearbit.com/' + value[12:]
                        image = download_image(image_url)
                        if image:
                            image_file_name = 'tempFile' + str(c) + '.png' # Save as PNG or JPG
                            save_image_to_file(image, image_file_name)
                            img = Image(image_file_name)
                            ws.add_image(img, logoDestination[c])
                            print(f"Image has been inserted into excel_file_name successfully.")
            else:
                value = 'N/A'
            ws.cell(counter,c + 2,value)
            ws.cell(counter,1,element)

        # Get earnings from 'symbol/earning' path of Yahoo Finance API
        i = 0
        url = "https://yahoo-finance-api-data.p.rapidapi.com/symbol/earning?symbol=" + symbols[c] + "&limit=10" 
        response = g(url,apiHost)
        quarterlyDate = [1,2,3,4,5]
        quarterlyRevenue = [1,2,3,4,5]
        quarterlyEarnings = [1,2,3,4,5]

        # go thru 4 quarters
        numKeys = len(response['data']['financialsChart']['quarterly'])
        for i in range(numKeys):
            if response['success'] == True:
                quarterlyDate[i] = response['data']['financialsChart']['quarterly'][i]['date']
                quarterlyRevenue[i] = response['data']['financialsChart']['quarterly'][i]['revenue']['fmt']
                quarterlyEarnings[i] = response['data']['financialsChart']['quarterly'][i]['earnings']['fmt']
            else:
                quarterlyDate[i] = 'N/A'
                quarterlyRevenue[i] = 'N/A'
                quarterlyEarnings[i] = 'N/A'

        # Output results for column 1 (if c==0)
        if c == 0:
            ws.cell(row=2, column=1, value="Name")
            ws.cell(row=3, column=1, value="Symbol")
            ws.cell(row=4, column=1, value="Revenue for " + quarterlyDate[0])
            ws.cell(row=5, column=1, value="Revenue for " + quarterlyDate[1])
            ws.cell(row=6, column=1, value="Revenue for " + quarterlyDate[2])
            ws.cell(row=7, column=1, value="Revenue for " + quarterlyDate[3])
            ws.cell(row=8, column=1, value="Earnings for " + quarterlyDate[0])
            ws.cell(row=9, column=1, value="Earnings for " + quarterlyDate[1])
            ws.cell(row=10, column=1, value="Earnings for " + quarterlyDate[2])
            ws.cell(row=11, column=1, value="Earnings for " + quarterlyDate[3])

        # Output results for column 2
        ws.cell(row=3, column=2+c, value=symbols[c])
        ws.cell(row=4, column=2+c, value=quarterlyRevenue[0])
        ws.cell(row=5, column=2+c, value=quarterlyRevenue[1])
        ws.cell(row=6, column=2+c, value=quarterlyRevenue[2])
        ws.cell(row=7, column=2+c, value=quarterlyRevenue[3])
        ws.cell(row=8, column=2+c, value=quarterlyEarnings[0])
        ws.cell(row=9, column=2+c, value=quarterlyEarnings[1])
        ws.cell(row=10, column=2+c, value=quarterlyEarnings[2])
        ws.cell(row=11, column=2+c, value=quarterlyEarnings[3])

        # Create a new sheet for each symbol
        new_sheet = wb.create_sheet(title=symbols[c])
        timeSpan = timePeriod
        loaded = getArray(symbols[c],timePeriod)     # 1d, 5d, 1mo, 3mo, 6mo, 1y, 2y, 5y, 10y, ytd, max
        if loaded[2] == 0:
            continue

        # Paste time stamps  
        timestamps = loaded[0]
        index = 1
        for i in timestamps:
            index += 1
            new_sheet.cell(row=index, column=1, value=i)

        # Paste values
        values = loaded[1]
        index = 1
        for i in values:
            index += 1
            new_sheet.cell(row=index, column=2, value= i)

        # Find maxrows
        maxRows = loaded[2]

        # Headers
        new_sheet.cell(1,1,timeSpan)
        new_sheet.cell(1,2,"")

        # Create area chart
        chart = AreaChart()

        # Cosmetic
        chart.title = symbols[c]
        chart.x_axis.title = timeSpan
        chart.y_axis.title = ""
        chart.x_axis.majorTickMark = "out"  # Display tick marks on x-axis
        chart.legend = None
        chart.y_axis.majorGridlines = None
        chart.width = 3.5
        chart.height = 3.5

        # References
        values = Reference(new_sheet, min_col=2, min_row=2, max_row=maxRows)
        x_values = Reference(new_sheet, min_col=1,min_row=2, max_row=maxRows)

        #time_reference = Reference(new_sheet, min_col=1,min_row=2,max_row=196)
        chart.add_data(values, titles_from_data = False)
        chart.set_categories(x_values)
        ws.add_chart(chart,stockDestination[c])

    symbolString = ''
    for symbol in symbols:
        symbolString += symbol + '%2C'
    symbolString = symbolString[:-3]
    url = "https://yahoo-finance-api-data.p.rapidapi.com/symbol/composite?symbol=" + symbolString
    response = g(url,apiHost)
    #print('last call = ' + response['data'][0]['shortName'])
    for j in range(numNonBlank):
        if validCode[j] == False:
            continue
        #print('j=' + str(j) + str(response['data']))
        ws.cell(2,j+2).value = response['data'][j]['shortName']
        ws.cell(33,j+2).value = response['data'][j]['currency']
        ws.cell(34,j+2).value = response['data'][j]['pegRatio']
        ws.cell(35,j+2).value = f"${(response['data'][j]['revenue']):,.2f}"
        ws.cell(36,j+2).value = response['data'][j]['epsCurrentYear']
        ws.cell(37,j+2).value = response['data'][j]['quoteSummary']['summaryDetail']['regularMarketPreviousClose']
        ws.cell(38,j+2).value = response['data'][j]['quoteSummary']['summaryDetail']['payoutRatio']

    # Make labels in column 1
        ws.cell(33,1).value = "currency"
        ws.cell(34,1).value = "pegRatio"
        ws.cell(35,1).value = "revenue"
        ws.cell(36,1).value = "epsCurrentYear"
        ws.cell(37,1).value = "regularMarketPreviousClose"
        ws.cell(38,1).value = "payoutRatio"

    # Make last cosmetic changes
    for i in range(c+2):
        if i % 2 == 0:
            columnColor = "FFFFFF"
        else:
            columnColor = "F2F2F2"
        for j in range(50):
            ws.cell(j+1,i+1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(j+1,i+1).fill = PatternFill(start_color = columnColor, end_color = columnColor, fill_type = "solid")
        for cell in ws[19]:
            if ws.cell(row=19,column = i+1).value != None and i > 0:
                tempLink = ws.cell(row = 19, column = i+1).value
                ws.cell(row = 19, column = i+1).hyperlink = tempLink
                ws.cell(row = 19, column = i+1).style = "Hyperlink"
                ws.cell(row = 19, column = i+1).fill = PatternFill(start_color = columnColor, end_color = columnColor, fill_type = "solid")
    for cell in ws[2]:
        cell.alignment = Alignment(wrap_text=True, horizontal = 'center', vertical = 'top')
    for cell in ws[13]:
        cell.alignment = Alignment(wrap_text=True, horizontal = 'center', vertical = 'top')

    # Now try this to let user download:
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

